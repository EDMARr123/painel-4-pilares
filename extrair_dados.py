r"""
Extrai os dados do painel "4 Pilares" a partir de SOMA NAO SALVA ENCIMA.xlsx
(aba "SOMAR 4 PILARES") e salva um dados.json pronto pro gerador de HTML
consumir.

Layout da aba (colunas, confirmado em 19/08 após reorganização da planilha
pelo Edmar — colunas de pilar/recompra/SKU/prêmio mudaram de lugar em
relação à versão de 17/08): blocos por supervisor (linha com nome do
supervisor na col D + cabeçalho "POSITIVAÇÃO" na col E), seguidos de 1
linha por RCA até a próxima linha de subtotal.

Por RCA:
- C=código, D=nome
- E/F/G/H    = positivação meta/real/falta/%
- J/K/L      = margem meta/real/%
- N/O/P      = mix meta/real/%
- R/S/T      = financeiro meta/real/falta R$; U = meta do dia
- W           = tendência % de fechamento (V = tendência R$, já bate com o
  projetado calculado abaixo)
- Z          = nº de pilares atingidos (0-4)
- AC/AD/AE/AF = industrializado meta/realizado/participação/margem % (pode ser negativa)
- AJ/AK/AL/AM = thermoprocessado meta/realizado/participação/margem % (pode ser negativa)
- BA/BB      = recompra (contagem/%) — espelha AO/AP, mas é o bloco com o
  cabeçalho "RECOMPRA" explícito
- BD         = média de pedidos
- BF/BG      = SKU meta/realizado
- BI         = prêmio R$ do industrializado
- BK         = prêmio R$ do thermoprocessado
- AS/AT      = positivação dia 15 (resultado / 1º prêmio R$)
- AV/AX      = positivação dia 30 (resultado / 2º prêmio R$)

Tendência "projetado R$" = REAL(S) / TRABALHADOS * DIAS_UTEIS (globais no
topo da planilha, linhas 4-5).
"""

import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile

import openpyxl

_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"

CAMINHO_SOMA = r"C:\Users\edmar\Desktop\CONTAR 4 PILARES\SOMA NAO SALVA ENCIMA.xlsx"

PASTA_BASE = os.path.dirname(os.path.abspath(__file__))
CAMINHO_SAIDA = os.path.join(PASTA_BASE, "dados.json")
CAMINHO_SAIDA_TOTAIS = os.path.join(PASTA_BASE, "totais_gerais.json")


def _num(v):
    """Blinda contra células com erro (#N/A etc) — acontece quando a planilha
    tem vínculos externos quebrados (ex: aviso "Não foi possível obter
    valores atualizados de uma pasta de trabalho vinculada" no Excel).
    Cai pra 0 em vez de derrubar a geração inteira do painel."""
    return v if isinstance(v, (int, float)) else 0


def _achar_external_link_xml(caminho_xlsx, contem_no_nome):
    """Acha o cache <externalLink> (xl/externalLinks/externalLinkN.xml) cujo
    arquivo de origem contém `contem_no_nome` no caminho, e devolve o XML já
    parseado — ou None se o vínculo não existir/tiver sido removido."""
    with zipfile.ZipFile(caminho_xlsx) as z:
        rels = [n for n in z.namelist() if re.fullmatch(r"xl/externalLinks/_rels/externalLink\d+\.xml\.rels", n)]
        for rels_path in rels:
            rels_xml = ET.fromstring(z.read(rels_path))
            if any(contem_no_nome.lower() in (rel.get("Target") or "").lower() for rel in rels_xml):
                link_path = rels_path.replace("_rels/", "").rsplit(".rels", 1)[0]
                return ET.fromstring(z.read(link_path))
    return None


def _cache_por_nome(external_link_el):
    """Lê o snapshot cacheado do <externalLink> e devolve {NOME (antes do
    primeiro ' - '): {coluna: valor}}. Casa por NOME em vez de código porque
    o THERMOPROCESSADO.xls usa um sistema de códigos diferente do resto da
    planilha (18/08: código do WESLEY é 15 na SOMA e 117 no export de
    Thermo — mesmo RCA, códigos diferentes — o que quebra o VLOOKUP por
    código e faz o realizado de Thermo cair pra 0 em todo mundo)."""
    if external_link_el is None:
        return {}

    def valor(cell):
        v = cell.find(f"{_NS}v")
        if v is None:
            return None
        if cell.get("t") == "str":
            return v.text
        try:
            return float(v.text)
        except (TypeError, ValueError):
            return v.text

    cache = {}
    for row in external_link_el.findall(f".//{_NS}sheetData/{_NS}row"):
        celulas = {re.match(r"^[A-Z]+", c.get("r")).group(): valor(c) for c in row}
        nome_bruto = celulas.get("B")
        if not isinstance(nome_bruto, str):
            continue
        nome = re.split(r"\s+-\s+", nome_bruto.strip(), maxsplit=1)[0].strip().upper()
        cache[nome] = celulas
    return cache


def _achar_no_cache(cache, nome_rca):
    """Casa `nome_rca` com uma chave do cache por igualdade ou prefixo (nos
    dois sentidos) — necessário porque alguns nomes vêm truncados na própria
    célula de origem da SOMA (ex: '471 - ANTONIO EDVALD', sem sobrenome)."""
    alvo = nome_rca.strip().upper()
    if alvo in cache:
        return cache[alvo]
    candidatos = [k for k in cache if k.startswith(alvo) or alvo.startswith(k)]
    if len(candidatos) == 1:
        return cache[candidatos[0]]
    return None


def extrair_totais(ws):
    """Bloco de totais gerais da planilha (linhas 82-98, coluna R = rótulo,
    T/U = meta/realizado) — Margem e Mix aqui são o número final calculado
    pelo Edmar na planilha, não uma média/soma das linhas por RCA.

    Conta-Corrente (linhas 109-114, adicionado 24/08): R=meta, S=realizado,
    T=tendência R$, V=%."""
    return {
        "margem": {"meta": _num(ws["T84"].value), "real": _num(ws["U84"].value)},
        "mix": {"meta": _num(ws["T87"].value), "real": _num(ws["U87"].value)},
        "meta_clientes": _num(ws["T92"].value),
        "realizado_clientes": _num(ws["T94"].value),
        "nao_comprou": _num(ws["T96"].value),
        "recompra_pct": _num(ws["T98"].value),
        "conta_corrente": {
            "meta": _num(ws["R114"].value),
            "realizado": _num(ws["S114"].value),
            "tendencia": _num(ws["T114"].value),
            "pct": _num(ws["V114"].value),
        },
    }


def extrair():
    wb = openpyxl.load_workbook(CAMINHO_SOMA, data_only=True)
    ws = wb["SOMAR 4 PILARES"]

    dias_uteis = ws.cell(row=4, column=6).value
    trabalhados = ws.cell(row=5, column=6).value

    # Bypassa o VLOOKUP quebrado de Thermo (veja _cache_por_nome) lendo o
    # realizado/margem direto do snapshot cacheado do vínculo externo.
    cache_thermo = _cache_por_nome(_achar_external_link_xml(CAMINHO_SOMA, "THERMOPROCESSADO"))

    rcas = []
    supervisor_atual = None
    for r in range(1, ws.max_row + 1):
        col_d = ws.cell(row=r, column=4).value
        col_e = ws.cell(row=r, column=5).value

        # Linha de cabeçalho de um novo bloco de supervisor.
        if col_d and col_e == "POSITIVAÇÃO":
            supervisor_atual = str(col_d).strip()
            continue

        codigo = ws.cell(row=r, column=3).value
        nome_bruto = col_d
        if supervisor_atual is None or codigo is None or nome_bruto is None:
            continue
        # Linha de subtotal do bloco (código vazio, mas nome preenchido) — pula.
        if str(codigo).strip() == "":
            continue

        codigo_str = str(int(float(str(codigo).strip())))
        # Nome vem como "23 - FABIO L. - GYN" — tira o prefixo de código
        # repetido; o que sobra depois do último " - " vira a "rota/praça"
        # exibida no card (ex: "GYN", "GYN RT 96 - HIDROLANDIA").
        nome_sem_codigo = str(nome_bruto).strip()
        prefixo = f"{codigo_str} - "
        if nome_sem_codigo.startswith(prefixo):
            nome_sem_codigo = nome_sem_codigo[len(prefixo):]
        if " - " in nome_sem_codigo:
            nome_rca, rota = nome_sem_codigo.rsplit(" - ", 1)
        else:
            nome_rca, rota = nome_sem_codigo, ""

        def val(col):
            v = ws.cell(row=r, column=col).value
            return v if isinstance(v, (int, float)) else 0

        real_financeiro = val(19)  # S
        meta_financeiro = val(18)  # R
        projetado = (real_financeiro / trabalhados * dias_uteis) if trabalhados else 0
        tendencia_pct = val(23)  # W = "TENTÊNCIA" (%) — bate com projetado/meta
        # U = "META DIA": quanto falta vender por dia útil restante pra bater a
        # meta do mês (vem negativo na planilha — é falta, não excedente).
        meta_dia = abs(val(21))
        # A coluna de % financeiro (real/meta) que existia antes sumiu na
        # reorganização de 19/08 — calcula direto em vez de depender de uma
        # coluna fixa.
        financeiro_pct = (real_financeiro / meta_financeiro) if meta_financeiro else 0

        # Layout confirmado em 19/08: AC/AD/AE/AF = meta/real/participação/
        # margem (industrializado); AJ/AK/AL/AM = idem (thermo); Z = pilar;
        # BA/BB = recompra; BD = média pedidos; BF/BG = SKU meta/real;
        # BI/BK = prêmio industrializado/thermo.
        industrializado_real = val(30)

        info_thermo = _achar_no_cache(cache_thermo, nome_rca)
        thermo_real = info_thermo.get("K") or 0 if info_thermo is not None else val(37)
        thermo_participacao_pct = (thermo_real / real_financeiro) if real_financeiro else 0
        # Margem % de Thermo depende do mesmo VLOOKUP por código quebrado (AI14);
        # a coluna O do cache de nome (usada pra bypassar o "real") não tem o
        # mesmo significado de margem que tem no arquivo de Industrializado —
        # em vez de arriscar mostrar um número inventado, mantém 0 até o
        # export do THERMOPROCESSADO.xls trazer o código certo na coluna B.
        thermo_margem_pct = val(39) if thermo_real else 0
        if thermo_margem_pct < 0:
            thermo_margem_pct = 0

        rcas.append({
            "codigo": codigo_str,
            "nome": nome_rca,
            "rota": rota,
            "supervisor": supervisor_atual,
            "pilares": {
                "positivacao": {"meta": val(5), "real": val(6), "pct": val(8)},
                "margem": {"meta": val(10), "real": val(11), "pct": val(12)},
                "mix": {"meta": val(14), "real": val(15), "pct": val(16)},
                "financeiro": {"meta": meta_financeiro, "real": real_financeiro, "pct": financeiro_pct},
            },
            "pilares_atingidos": int(val(26)),
            "tendencia": {"pct": tendencia_pct, "projetado": projetado, "meta": meta_financeiro, "meta_dia": meta_dia},
            "industrializado": {"meta": val(29), "real": industrializado_real, "participacao_pct": val(31), "margem_pct": val(32), "premio": val(61)},
            "thermo": {"meta": val(36), "real": thermo_real, "participacao_pct": thermo_participacao_pct, "margem_pct": thermo_margem_pct, "premio": val(63)},
            "recompra_pct": val(54),  # BB = "RECOMPRA" %
            "media_pedidos": val(56),  # BD = "MÉDIA PEDIDOS"
            "sku": {"meta": val(58), "real": val(59)},  # BF/BG = "SKU" meta/realizado
            "positivacao_dia15": {"resultado": val(45), "premio": val(46)},  # AS/AT
            "positivacao_dia30": {"resultado": val(48), "premio": val(50)},  # AV/AX
        })

    return rcas


def main():
    rcas = extrair()
    with open(CAMINHO_SAIDA, "w", encoding="utf-8") as f:
        json.dump(rcas, f, ensure_ascii=False, indent=2)
    print(f"{len(rcas)} RCAs extraídos. Salvo em: {CAMINHO_SAIDA}")

    wb = openpyxl.load_workbook(CAMINHO_SOMA, data_only=True)
    totais = extrair_totais(wb["SOMAR 4 PILARES"])
    with open(CAMINHO_SAIDA_TOTAIS, "w", encoding="utf-8") as f:
        json.dump(totais, f, ensure_ascii=False, indent=2)
    print(f"Totais gerais salvos em: {CAMINHO_SAIDA_TOTAIS}")


if __name__ == "__main__":
    main()
